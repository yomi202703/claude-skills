#!/usr/bin/env python3
"""作業報告書テンプレート(template.xlsx)に値を差し込んで .xlsx を出力する。

本文(作業内容)の生成は行わない。呼び出し側(LLM)が生成した本文と、
日付・作業時間などのフィールドを JSON で受け取り、テンプレートの該当セルに
流し込むだけの決定論的な処理。

使い方:
    python3 fill_report.py --json data.json
    python3 fill_report.py --json -        # 標準入力から JSON を読む

JSON フィールド (すべて任意。未指定はテンプレート既定/空欄):
    company        提出先会社名 (既定 "InfoDeliver")  -> A2
    name           氏名         (既定 "三浦由人")      -> G4 "氏名：…"
    subject        件名         (既定 "インターン")    -> C13
    work_date      作業日 "YYYY-MM-DD" (必須)          -> C14
    time_start     開始 "HH:MM" (必須)                 -> C15
    time_end       終了 "HH:MM" (必須)                 -> C15
    break_min      休憩(分) 整数 (既定 0)              -> C15
    body           作業内容 本文 (必須)                -> C16
    special_notes  特記事項 (既定 空)                  -> C27
    progress       作業進捗状況 (既定 "途中")          -> C34
    next_plan_date 次回作業予定日 "YYYY-MM-DD" (任意)  -> C35
    next_plan      次回の作業予定 (既定 "続き")        -> C36
    out            出力先パス (任意。既定 ./作業報告書_YYYYMMDD.xlsx)

終了コード: 成功 0 / 入力不備・テンプレート不在 1。
"""
import argparse
import json
import sys
from pathlib import Path

TEMPLATE = Path(__file__).resolve().parent / "template.xlsx"

DEFAULTS = {
    "company": "InfoDeliver",
    "name": "三浦由人",
    "subject": "インターン",
    "break_min": 0,
    "special_notes": "",
    "progress": "途中",
    "next_plan": "続き",
}


def _ymd(s):
    """'YYYY-MM-DD' -> (year, month, day) int 三つ組。不正なら ValueError。"""
    y, m, d = s.split("-")
    return int(y), int(m), int(d)


def _fmt_date(s):
    y, m, d = _ymd(s)
    return f"　{y} 年 {m} 月 {d} 日"


def _fmt_time(start, end, brk):
    return f"{start} ～ {end}　（うち休憩 {brk} 分）"


def _fmt_next(s):
    if not s:
        return "　年　月　日　　：　～　："
    y, m, d = _ymd(s)
    return f"　{y} 年 {m} 月 {d} 日　　：　～　："


def build(data):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl が必要です: pip install openpyxl")

    if not TEMPLATE.exists():
        sys.exit(f"テンプレートがありません: {TEMPLATE}")

    cfg = {**DEFAULTS, **{k: v for k, v in data.items() if v is not None}}

    for req in ("work_date", "time_start", "time_end", "body"):
        if not cfg.get(req):
            sys.exit(f"必須フィールドが未指定: {req}")

    try:
        _ymd(cfg["work_date"])
    except (ValueError, AttributeError):
        sys.exit(f"work_date は YYYY-MM-DD 形式で: {cfg['work_date']!r}")

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active
    if ws is None:
        sys.exit("テンプレートにアクティブシートがありません")

    ws["A2"] = cfg["company"]
    ws["G4"] = f"氏名：{cfg['name']}"
    ws["C13"] = cfg["subject"]
    ws["C14"] = _fmt_date(cfg["work_date"])
    ws["C15"] = _fmt_time(cfg["time_start"], cfg["time_end"], cfg["break_min"])
    ws["C16"] = cfg["body"]
    ws["C27"] = cfg["special_notes"]
    ws["C34"] = cfg["progress"]
    ws["C35"] = _fmt_next(cfg.get("next_plan_date"))
    ws["C36"] = cfg["next_plan"]

    out = cfg.get("out")
    if not out:
        ymd = cfg["work_date"].replace("-", "")
        out = f"作業報告書_{ymd}.xlsx"
    out = Path(out).expanduser()
    wb.save(out)
    return out


def main():
    ap = argparse.ArgumentParser(description="作業報告書テンプレートへの差し込み")
    ap.add_argument("--json", required=True, help="JSON ファイルパス、または - で標準入力")
    args = ap.parse_args()

    raw = sys.stdin.read() if args.json == "-" else Path(args.json).read_text(encoding="utf-8")
    data = json.loads(raw)
    out = build(data)
    print(f"出力: {out}")


if __name__ == "__main__":
    main()
