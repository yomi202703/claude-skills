"""Deterministic golden regression for xlsx_to_html.py.

The HTML converter is fully deterministic, so a golden fingerprint per sheet
guards its output (per REDESIGN_2026-06.md "deterministic golden survives").

Fingerprint = faithfulness (must be 100%), rowspan/colspan count, <tr> count,
and a short sha256 of the HTML (catches any byte-level drift). We store the
fingerprint rather than the full HTML to keep goldens small and readable.

On first run (or intentional converter changes): delete the corresponding
<basename>.yml under tests/test_html_regression/ and re-run; pytest-regressions
writes fresh baselines and fails once so you can inspect + commit them.

Run: pytest ~/.claude/skills/xlsx-router/_dev/tests/ -q
"""
import hashlib
import re
import sys
from pathlib import Path

import openpyxl
import pytest

SKILL = Path.home() / ".claude/skills/xlsx-router"
sys.path.insert(0, str(SKILL / "scripts"))
import xlsx_to_html as X  # noqa: E402

CORPUS = sorted((SKILL / "_dev/corpus").glob("*.xlsx"))


def _safe_basename(path: Path) -> str:
    stem = re.sub(r"[【】（）・]", "_", path.stem)
    return "html_" + re.sub(r"_+", "_", stem).strip("_")


# Sheets above this cell budget route to the SQLite path (triage "huge" exception),
# not HTML — so the HTML converter golden only covers HTML-path sheets. This also
# keeps the hook-fired suite fast (huge-sheet verify dominates runtime).
HTML_PATH_CELL_BUDGET = 20_000


def _fingerprint(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row * ws.max_column > HTML_PATH_CELL_BUDGET:
            out[sn] = {"skipped": "huge→SQLite path"}
            continue
        anno = X._drawing_annotations(path, sn)
        html = X.sheet_to_html(ws, anno)
        present, total, _ = X.verify_faithful(ws, html)
        out[sn] = {
            "faithful": f"{present}/{total}",
            "spans": len(re.findall(r"(?:rowspan|colspan)=", html)),
            "tr": html.count("<tr>"),
            "sha256_16": hashlib.sha256(html.encode("utf-8")).hexdigest()[:16],
        }
    wb.close()
    return out


@pytest.mark.parametrize("xlsx", CORPUS, ids=lambda p: p.stem)
def test_html_faithful_and_deterministic(xlsx: Path, data_regression):
    fp = _fingerprint(xlsx)
    # Hard invariant independent of the golden: every HTML-path sheet 100% faithful.
    for sn, d in fp.items():
        if "faithful" not in d:  # huge→SQLite sheet, not converted to HTML
            continue
        present, total = d["faithful"].split("/")
        assert present == total, f"{xlsx.name}/{sn}: not 100% faithful ({d['faithful']})"
    data_regression.check(fp, basename=_safe_basename(xlsx))
