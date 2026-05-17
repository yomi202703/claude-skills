#!/usr/bin/env python3
"""Self-improve用のテストコーパスを生成する。

各シートケース(P1-P5)をカバーする6つのxlsxを作る。
再実行で再生成される（idempotent）。
"""
from pathlib import Path
import openpyxl

CORPUS = Path(__file__).parent / "corpus"
CORPUS.mkdir(exist_ok=True)


def db_small():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "orders"
    ws.append(["id", "name", "qty", "price", "date"])
    for i in range(1, 21):
        ws.append([i, f"product_{i}", i * 2, i * 100, "2026-01-01"])
    wb.save(CORPUS / "db_small.xlsx")


def db_large():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "log"
    ws.append([f"col_{i}" for i in range(1, 11)])
    for r in range(500):
        ws.append([f"v{r}_{c}" for c in range(10)])
    wb.save(CORPUS / "db_large.xlsx")


def structured_small():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "チェックシート"
    ws["A1"] = "【法人契約チェックシート】"
    ws.merge_cells("A1:E1")
    ws["A3"] = "区分"; ws["B3"] = "帳票"; ws["C3"] = "項目"; ws["D3"] = "確認事項"; ws["E3"] = "チェック"
    ws["A4"] = "新契約"
    ws.merge_cells("A4:A30")
    ws["B4"] = "申込書"
    ws.merge_cells("B4:B15")
    ws["B16"] = "取扱報告書"
    ws.merge_cells("B16:B30")
    for i in range(4, 31):
        ws[f"C{i}"] = f"項目{i-3}"
        ws[f"D{i}"] = f"確認事項の内容 {i-3}"
        ws[f"E{i}"] = "はい/いいえ"
    wb.save(CORPUS / "structured_small.xlsx")


def structured_large():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "大規模ルール"
    ws["A1"] = "【大規模チェックシート】"
    ws.merge_cells("A1:E1")
    ws["A3"] = "区分"; ws["B3"] = "帳票"; ws["C3"] = "項目"; ws["D3"] = "確認事項"; ws["E3"] = "チェック"
    ws["A4"] = "新契約"
    ws.merge_cells("A4:A120")
    ws["A121"] = "増額"
    ws.merge_cells("A121:A203")
    forms = ["申込書", "意向確認書", "取扱報告書", "届出書", "同意書"]
    row = 4
    for form in forms:
        span = 40
        ws[f"B{row}"] = form
        ws.merge_cells(f"B{row}:B{row+span-1}")
        row += span
    for i in range(4, 204):
        ws[f"C{i}"] = f"項目{i-3}"
        ws[f"D{i}"] = f"確認事項の内容 {i-3}（詳細は特記事項参照のこと）"
        ws[f"E{i}"] = "はい/非該当/いいえ"
    wb.save(CORPUS / "structured_large.xlsx")


def multi_sheet():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("basic_info")
    ws1.append(["id", "name", "type"])
    for i in range(10):
        ws1.append([i, f"item{i}", "A"])
    ws2 = wb.create_sheet("master_table")
    ws2.append(["code", "label", "cat"])
    for i in range(2000):
        ws2.append([f"M{i:04d}", f"label{i}", "cat1"])
    ws3 = wb.create_sheet("rules_sheet")
    ws3["A1"] = "【ルール一覧】"
    ws3.merge_cells("A1:C1")
    ws3["A2"] = "区分"; ws3["B2"] = "内容"; ws3["C2"] = "判定"
    ws3["A3"] = "新規"
    ws3.merge_cells("A3:A20")
    for i in range(3, 21):
        ws3[f"B{i}"] = f"内容{i}"
        ws3[f"C{i}"] = "OK/NG"
    wb.save(CORPUS / "multi_sheet.xlsx")


def huge():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "big_data"
    ws.append([f"col_{i}" for i in range(1, 51)])
    for r in range(2000):
        ws.append([f"v{r}_{c}" for c in range(50)])
    wb.save(CORPUS / "huge.xlsx")


if __name__ == "__main__":
    db_small()
    db_large()
    structured_small()
    structured_large()
    multi_sheet()
    huge()
    print(f"Generated 6 corpus files in {CORPUS}")
