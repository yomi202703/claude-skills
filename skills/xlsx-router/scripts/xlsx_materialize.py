#!/usr/bin/env python3
"""xlsxをSQLiteに一括変換（Claude Code用 取り込みスクリプト）。

SQLite 経路は巨大・均一な db ライクシート専用（複雑シートは HTML 経路）。
ヘッダーは先頭の非空行、データはその下〜末尾（末尾空行トリム）とする。
結合セルの anchor 値はヘッダー・データ行いずれも forward-fill する。

使用例:
  python3 xlsx_materialize.py foo.xlsx
  python3 xlsx_materialize.py foo.xlsx --out data/foo.db
  python3 xlsx_materialize.py foo.xlsx --tables "003申込書,005取扱者報告書"
  python3 xlsx_materialize.py foo.xlsx --skip-merged
"""
import argparse
import sqlite3
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import xlsx_primitives as _prim  # noqa: E402
from xlsx_drawings import extract as _extract_drawings  # noqa: E402


def _simple_header_and_data(ws):
    """Header = first non-empty row; data = rows beneath it (trailing blanks
    trimmed). The SQLite path only receives large, uniform db-like sheets —
    complex/merged sheets take the HTML path — so a row-1-style header is correct
    here and no heuristic header scoring is needed.
    """
    first = last = None
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v is not None and (not isinstance(v, str) or v.strip()) for v in row):
            if first is None:
                first = r
            last = r
    if first is None:
        return [1], 2, 1
    return [first], first + 1, (last if last is not None else first)


def _write_drawings_table(cur, sheet_slug: str, shapes: list) -> int:
    """Create <sheet_slug>_drawings table and insert every shape/pic/cxn row.

    Returns number of rows inserted. Each shape row preserves the anchor
    cellref so downstream can JOIN drawings to the tabular data when needed.
    """
    table = _prim.sqlite_column_name(sheet_slug, "sheet") + "_drawings"
    cur.execute(f'DROP TABLE IF EXISTS "{table}"')
    cur.execute(
        f'CREATE TABLE "{table}" ('
        '"kind" TEXT, "anchor_range" TEXT, "anchor_from" TEXT, "anchor_to" TEXT, '
        '"name" TEXT, "text" TEXT, "media_path" TEXT, "extracted_path" TEXT'
        ')'
    )
    rows = 0
    for s in shapes:
        cur.execute(
            f'INSERT INTO "{table}" VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (
                s.get("kind"),
                s.get("anchor_range"),
                s.get("anchor_from"),
                s.get("anchor_to"),
                s.get("name"),
                s.get("text") or "",
                s.get("media_path"),
                s.get("extracted_path"),
            ),
        )
        rows += 1
    return rows


def main():
    p = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    p.add_argument("file")
    p.add_argument("--out", help="output SQLite path (default: <file>.sqlite)")
    p.add_argument(
        "--tables", help="comma-separated sheet names to include (default: all)"
    )
    p.add_argument(
        "--skip-merged", action="store_true", help="skip sheets with merged cells"
    )
    p.add_argument(
        "--header-rows",
        help="comma-separated header row indices (e.g. '22' or '9,10') when the "
        "real header is NOT the first non-empty row (spec block above the table). "
        "Supply this from your own read of the top rows; default: first non-empty row.",
    )
    args = p.parse_args()

    out_path = args.out or str(Path(args.file).with_suffix(".sqlite"))
    wanted = set(s.strip() for s in args.tables.split(",")) if args.tables else None

    wb = openpyxl.load_workbook(args.file, data_only=True)
    conn = sqlite3.connect(out_path)
    cur = conn.cursor()

    # Drawings: extract once, stage media alongside the SQLite output.
    drawings_dir = Path(out_path).parent / (Path(out_path).stem + "_drawings")
    try:
        drawings_all = _extract_drawings(Path(args.file).resolve(), extract_dir=drawings_dir)
        drawings_by_sheet = drawings_all.get("sheets", {})
    except Exception as e:
        print(f"[warn] drawings extraction failed: {e}", file=sys.stderr)
        drawings_by_sheet = {}

    for sn in wb.sheetnames:
        if wanted and sn not in wanted:
            continue
        ws = wb[sn]
        merged = len(list(ws.merged_cells.ranges))
        if merged > 0 and args.skip_merged:
            print(f"[skip] {sn}: {merged} merged ranges", file=sys.stderr)
            continue
        if merged > 0:
            print(
                f"[warn] {sn}: {merged} merged ranges — forward-filled header and data",
                file=sys.stderr,
            )

        sheet_draw = drawings_by_sheet.get(sn) or {}
        # Capture floating shape/pic annotations regardless — queryable alongside data.
        shapes = sheet_draw.get("shapes") or []
        sheet_slug = _prim.sanitize_slug(sn)
        if shapes:
            draw_rows = _write_drawings_table(cur, sheet_slug, shapes)
            print(
                f"[ok]   {sn} -> \"{sheet_slug}_drawings\" ({draw_rows} shapes/pics)",
                file=sys.stderr,
            )
        if args.header_rows:
            hr = [int(x) for x in args.header_rows.split(",") if x.strip()]
            _, _, data_end = _simple_header_and_data(ws)
            header_rows, data_start = hr, max(hr) + 1
        else:
            header_rows, data_start, data_end = _simple_header_and_data(ws)

        merge_lookup = _prim.build_merge_lookup(ws)
        labels = _prim.concat_header_labels(ws, header_rows, merge_lookup)
        unique_cols = _prim.uniquify(labels)

        table = _prim.sqlite_column_name(sn, "sheet")
        cols_def = ", ".join(f'"{c}" TEXT' for c in unique_cols)
        cur.execute(f'DROP TABLE IF EXISTS "{table}"')
        cur.execute(f'CREATE TABLE "{table}" ({cols_def})')

        placeholders = ",".join("?" * len(unique_cols))
        row_count = 0
        for row in _prim.iter_rows_ff(ws, data_start, data_end, merge_lookup):
            if not any(
                v is not None and (not isinstance(v, str) or v.strip()) for v in row
            ):
                continue
            vals = ["" if v is None else str(v) for v in row]
            vals = vals[: len(unique_cols)] + [""] * max(
                0, len(unique_cols) - len(vals)
            )
            cur.execute(f'INSERT INTO "{table}" VALUES ({placeholders})', vals)
            row_count += 1

        print(
            f'[ok]   {sn} -> "{table}" '
            f"(hdr=rows{header_rows}, {row_count} rows, {len(unique_cols)} cols)",
            file=sys.stderr,
        )

    conn.commit()
    conn.close()
    print(out_path)


if __name__ == "__main__":
    main()
