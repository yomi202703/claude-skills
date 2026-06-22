#!/usr/bin/env python3
"""xlsx シートを「忠実な構造保存 HTML」に変換する（Claude Code 用 既定成果物）。

設計: _dev/REDESIGN_2026-06.md。原典に忠実かつ AI ネイティブなテキスト表現を
1シート=1 HTML で出す。解釈は一切入れない（ヘッダを当てない・選択肢を分割しない・
forward-fill で結合を潰さない）。原典構造をそのまま写し、消費側 LLM が直読する。

忠実に写すもの:
  - 結合セル        → rowspan / colspan（無損失）
  - 静的なセル色/塗り → inline style（色=意味を保存）
  - 日付シリアル     → number_format を見て ISO 日付化（46163 問題の確定対処）
  - 図形/シェイプ文字 → アンカーセルへ注記埋め込み（xlsx_drawings 由来）

忠実性 hard ゲート: 変換後、原典の全非空セル文字列が HTML に存在するか自己照合する。
欠落があれば exit 2（--no-verify で無効化）。

使用例:
  python3 xlsx_to_html.py foo.xlsx                          # 全シート → <foo>_html/<sheet>.html
  python3 xlsx_to_html.py foo.xlsx --sheet "基本情報一覧"     # 1シート → stdout 経路にも対応
  python3 xlsx_to_html.py foo.xlsx --out-dir data/foo
  python3 xlsx_to_html.py foo.xlsx --sheet S --out data/foo/s.html
"""
import argparse
import datetime
import html as _html
import json
import re
import sys
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import xlsx_primitives as _prim  # noqa: E402

try:
    from xlsx_drawings import extract as _extract_drawings  # noqa: E402
except Exception:  # drawings extraction is best-effort
    _extract_drawings = None


_DATE_FMT_RE = re.compile(r"(yy|mm|dd|m/d|d/m|h:mm|mmm)", re.IGNORECASE)


def _date_aware(v, fmt):
    """Convert Excel date serials to ISO when number_format is a date format.

    Static cells formatted General come back as plain int even when they hold a
    serial; we ONLY convert when the cell's number_format is itself a date format
    (mirrors merges.md「日付シリアル値」). Never blanket-convert bare numbers.
    """
    if isinstance(v, datetime.datetime):
        return v.date().isoformat() if v.time() == datetime.time(0, 0) else v.isoformat(sep=" ")
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, (int, float)) and not isinstance(v, bool) and fmt and _DATE_FMT_RE.search(fmt):
        if 1 < v < 80000:
            try:
                return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))).isoformat()
            except Exception:
                return v
    return v


def _hex6(color):
    """Return 'RRGGBB' from an openpyxl Color if it carries a usable ARGB rgb, else None."""
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and len(rgb) == 8:
        return rgb[2:].upper()
    return None


def _style_attr(cell) -> str:
    parts = []
    try:
        if cell.font and cell.font.color is not None:
            h = _hex6(cell.font.color)
            if h and h != "000000":
                parts.append(f"color:#{h}")
    except Exception:
        pass
    try:
        fill = cell.fill
        if fill is not None and getattr(fill, "patternType", None) == "solid":
            h = _hex6(fill.fgColor)
            if h and h not in ("FFFFFF", "000000"):
                parts.append(f"background:#{h}")
    except Exception:
        pass
    if cell.font and getattr(cell.font, "bold", False):
        parts.append("font-weight:bold")
    return f' style="{";".join(parts)}"' if parts else ""


def _used_bounds(ws):
    """Last row/col carrying any non-empty cell (trims trailing blank padding).

    Uses iter_rows(values_only=True) — ~100× faster than per-cell ws.cell()
    access on large sheets.
    """
    last_r = last_c = 0
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for ci, v in enumerate(row, start=1):
            if v is not None and (not isinstance(v, str) or v.strip()):
                if r > last_r:
                    last_r = r
                if ci > last_c:
                    last_c = ci
    return max(1, last_r), max(1, last_c)


# Sheets above this used-cell budget route to the SQLite path (triage "huge"
# exception): their faithful HTML would overflow a downstream LLM's context, so
# we hand them to xlsx_materialize instead. The boundary is deterministic
# arithmetic — never a model judgment (per REDESIGN_2026-06.md triage).
HUGE_CELL_BUDGET = 20_000


def _extract_sheets(path: Path) -> dict:
    """Extract drawings once per workbook → {sheet_name: drawings_info}."""
    if _extract_drawings is None:
        return {}
    try:
        return _extract_drawings(path).get("sheets", {})
    except Exception:
        return {}


def _anno_from_shapes(shapes) -> dict:
    """Map (row, col) -> [shape texts] from a sheet's extracted shapes list."""
    anno: dict = {}
    for sh in shapes or []:
        txt = (sh.get("text") or "").strip()
        if not txt:
            continue
        rc = sh.get("anchor_from_rc")  # [row0, col0]
        if not rc:
            continue
        anno.setdefault((rc[0] + 1, rc[1] + 1), []).append(txt)
    return anno


def _drawing_annotations(path: Path, sheet: str) -> dict:
    """Convenience: extract + filter annotations for a single sheet (standalone use)."""
    return _anno_from_shapes(_extract_sheets(path).get(sheet, {}).get("shapes"))


def sheet_to_html(ws, annotations: Optional[dict] = None, bounds=None) -> str:
    annotations = annotations or {}
    last_r, last_c = bounds if bounds else _used_bounds(ws)

    span = {}
    skip = set()
    for mr in ws.merged_cells.ranges:
        span[(mr.min_row, mr.min_col)] = (mr.max_row - mr.min_row + 1, mr.max_col - mr.min_col + 1)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    skip.add((r, c))

    # Redirect annotations that land on a merged non-anchor cell to the anchor.
    anchor_of = {}
    for (ar, ac), (rs, cs) in span.items():
        for r in range(ar, ar + rs):
            for c in range(ac, ac + cs):
                anchor_of[(r, c)] = (ar, ac)
    redirected: dict = {}
    for (r, c), texts in annotations.items():
        tgt = anchor_of.get((r, c), (r, c))
        redirected.setdefault(tgt, []).extend(texts)

    trailing_notes = []  # shapes anchored outside the used grid
    rows_html = []
    # iter_rows(values_only=False) yields Cell objects (value + style) in one
    # efficient pass — far faster than ws.cell(r, c) per cell on large sheets.
    for row in ws.iter_rows(min_row=1, max_row=last_r, min_col=1, max_col=last_c):
        cells = []
        for cell in row:
            r, c = cell.row, cell.column
            if (r, c) in skip:
                continue
            v = _date_aware(cell.value, cell.number_format)
            txt = "" if v is None else _html.escape(str(v))
            note = ""
            if (r, c) in redirected:
                joined = " ／ ".join(redirected.pop((r, c)))
                note = f'<span class="anno">〔図形: {_html.escape(joined)}〕</span>'
            attrs = ""
            if (r, c) in span:
                rs, cs = span[(r, c)]
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'
            attrs += _style_attr(cell)
            cells.append(f"<td{attrs}>{txt}{note}</td>")
        rows_html.append("  <tr>" + "".join(cells) + "</tr>")

    # Any annotations not placed (anchored beyond used bounds) → trailing list.
    for (r, c), texts in redirected.items():
        trailing_notes.append(f"{_col_letter(c)}{r}: " + " ／ ".join(texts))

    parts = [f'<table data-sheet="{_html.escape(ws.title)}" data-rows="{last_r}" data-cols="{last_c}">']
    parts.extend(rows_html)
    parts.append("</table>")
    if trailing_notes:
        parts.append('<ul class="drawings">')
        parts.extend(f"  <li>〔図形 {_html.escape(n)}〕</li>" for n in trailing_notes)
        parts.append("</ul>")
    return "\n".join(parts) + "\n"


def _col_letter(idx1: int) -> str:
    s = ""
    n = idx1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(ord("A") + rem) + s
    return s


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", _html.unescape(s)).strip()


def verify_faithful(ws, html_text: str) -> tuple:
    """Return (present, total, misses_sample). Compares normalized source cell
    strings against the de-tagged HTML text. 100% = no source content dropped."""
    txt = _norm(re.sub(r"<[^>]+>", " ", html_text))
    total = present = 0
    misses = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = _norm(str(_date_aware(v, cell.number_format)))
            if not s:
                continue
            total += 1
            if s in txt:
                present += 1
            elif len(misses) < 10:
                misses.append((cell.row, cell.column, len(s)))
    return present, total, misses


def convert(path: Path, sheet: Optional[str], out: Optional[Path], out_dir: Optional[Path],
            do_verify: bool) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    names = [sheet] if sheet else list(wb.sheetnames)
    if sheet and sheet not in wb.sheetnames:
        sys.exit(f"[error] sheet '{sheet}' not found. available: {wb.sheetnames}")

    if out_dir is None and out is None:
        out_dir = path.parent / (path.stem + "_html")

    # Extract drawings once per workbook (feeds both annotations and image flags).
    drawings_sheets = _extract_sheets(path)

    results = []
    failed = False
    for sn in names:
        ws = wb[sn]
        last_r, last_c = _used_bounds(ws)
        cells = last_r * last_c
        draw = drawings_sheets.get(sn, {})
        counts = draw.get("counts") or {}
        n_draw = sum(int(v) for v in counts.values())

        # Triage (deterministic): huge sheets overflow context → SQLite path.
        if cells > HUGE_CELL_BUDGET:
            print(f"[huge] {sn}: {cells} cells > {HUGE_CELL_BUDGET} -> SQLite path "
                  f"(run xlsx_materialize.py)", file=sys.stderr)
            results.append({"sheet": sn, "path": "sqlite",
                            "reason": f"huge ({cells} cells)", "cells": cells,
                            "has_drawings": n_draw > 0})
            continue

        anno = _anno_from_shapes(draw.get("shapes"))
        html_text = sheet_to_html(ws, anno, bounds=(last_r, last_c))

        slug = _prim.sanitize_slug(sn)
        if out is not None and len(names) == 1:
            target = out
        else:
            target = (out_dir or path.parent) / f"{slug}.html"
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(html_text, encoding="utf-8")

        rec = {"sheet": sn, "path": "html", "html_path": str(target),
               "bytes": len(html_text.encode("utf-8")),
               "has_drawings": n_draw > 0,
               "shape_count": int(counts.get("shape", 0)),
               "pic_count": int(counts.get("pic", 0))}
        # Drawings present → flag as an image candidate; the consuming agent
        # decides whether the HTML (with shape annotations) suffices or a
        # rendered PNG is needed (layout-dependent meaning).
        if n_draw > 0:
            rec["suggests_image"] = True
        if do_verify:
            present, total, misses = verify_faithful(ws, html_text)
            rec["faithful"] = f"{present}/{total}"
            rec["faithful_pct"] = round(100 * present / total, 2) if total else 100.0
            if present < total:
                failed = True
                rec["miss_sample_rc_len"] = misses
                print(f"[FAIL] {sn}: {total - present} source cells missing from HTML "
                      f"(sample rc/len: {misses})", file=sys.stderr)
            else:
                print(f"[ok]   {sn} -> {target.name} ({total} cells, 100% faithful)",
                      file=sys.stderr)
        else:
            print(f"[ok]   {sn} -> {target.name}", file=sys.stderr)
        results.append(rec)

    out_obj = {"file": str(path), "sheets": results,
               "out_dir": str(out_dir) if out_dir else None}
    print(json.dumps(out_obj, ensure_ascii=False, indent=2))
    if failed:
        sys.exit(2)
    return out_obj


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("file")
    ap.add_argument("--sheet", help="single sheet name (default: all sheets)")
    ap.add_argument("--out", help="output html path (single sheet only)")
    ap.add_argument("--out-dir", help="output directory (default: <file>_html/)")
    ap.add_argument("--no-verify", action="store_true",
                    help="skip the faithfulness hard gate")
    args = ap.parse_args()
    convert(
        Path(args.file).resolve(),
        args.sheet,
        Path(args.out).resolve() if args.out else None,
        Path(args.out_dir).resolve() if args.out_dir else None,
        do_verify=not args.no_verify,
    )


if __name__ == "__main__":
    main()
