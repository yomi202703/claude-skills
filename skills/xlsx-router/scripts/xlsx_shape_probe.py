#!/usr/bin/env python3
"""Structural shape probe for xlsx sheets.

Emits a compact JSON description per sheet for downstream routing / debugging.
Never includes raw cell values — only lengths, types, and positions.

Usage:
    python3 xlsx_shape_probe.py <file.xlsx>            # → stdout JSON
    python3 xlsx_shape_probe.py <file.xlsx> --sheet N  # → single sheet
"""
import argparse
import json
import statistics
import sys
from collections import Counter
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import xlsx_primitives as _prim  # noqa: E402


FIRST_ROWS_LIMIT = 30
DEEP_SAMPLE_START = 30
DEEP_SAMPLE_END = 230
LONG_CELL_THRESHOLD = 150


def _merges_summary(ws) -> dict:
    count = 0
    max_rowspan = 0
    max_colspan = 0
    banner_rows = set()
    label_col_merges = 0
    for r in ws.merged_cells.ranges:
        count += 1
        rspan = r.max_row - r.min_row + 1
        cspan = r.max_col - r.min_col + 1
        max_rowspan = max(max_rowspan, rspan)
        max_colspan = max(max_colspan, cspan)
        if rspan == 1 and cspan >= max(3, ws.max_column - 1):
            banner_rows.add(r.min_row)
        if cspan <= 2 and rspan >= 3:
            label_col_merges += 1
    return {
        "count": count,
        "max_rowspan": max_rowspan,
        "max_colspan": max_colspan,
        "banner_rows": sorted(banner_rows),
        "label_col_merges": label_col_merges,
    }


def _probe_sheet(wb, sheet_name: str) -> dict:
    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    merges = _merges_summary(ws)
    merge_lookup = _prim.build_merge_lookup(ws)

    first_rows = []
    banner_set = set(merges["banner_rows"])
    v_merge_rows = set()
    for mr in ws.merged_cells.ranges:
        if mr.min_row != mr.max_row:
            for rr in range(mr.min_row, mr.max_row + 1):
                v_merge_rows.add(rr)

    fr_limit = min(FIRST_ROWS_LIMIT, max_row)
    for i in range(1, fr_limit + 1):
        lens = [_prim.cell_len(ws.cell(row=i, column=c).value) for c in range(1, max_col + 1)]
        widths_nb = sum(1 for L in lens if L > 0)
        eff_width = _prim.row_effective_width(ws, i, merge_lookup)
        flag = None
        if i in banner_set:
            flag = "banner"
        elif i in v_merge_rows:
            flag = "v-merge"
        first_rows.append({
            "idx": i,
            "widths_nb": widths_nb,
            "widths_eff": eff_width,
            "lens": lens,
            "merge_flag": flag,
        })

    long_cells_rows = []
    long_cells_counts = {}
    long_cells_max_len = 0
    long_scan_limit = min(FIRST_ROWS_LIMIT * 2, max_row)
    for i in range(1, long_scan_limit + 1):
        cnt_long = 0
        for c in range(1, max_col + 1):
            v = ws.cell(row=i, column=c).value
            if isinstance(v, str) and len(v) >= LONG_CELL_THRESHOLD:
                cnt_long += 1
                long_cells_max_len = max(long_cells_max_len, len(v))
        if cnt_long > 0:
            long_cells_rows.append(i)
            long_cells_counts[str(i)] = cnt_long

    widths = []
    sample_limit = min(DEEP_SAMPLE_END, max_row)
    for i in range(min(DEEP_SAMPLE_START, max_row) + 1, sample_limit + 1):
        widths.append(_prim.row_effective_width(ws, i, merge_lookup))
    median_width = int(statistics.median(widths)) if widths else 0
    hist = Counter(widths)
    top_hist = {str(k): v for k, v in hist.most_common(5)}

    scan_rows = min(200, max_row)
    total_cells = scan_rows * max_col
    nonempty = 0
    for i in range(1, scan_rows + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=i, column=c).value
            if v is not None and (not isinstance(v, str) or v.strip()):
                nonempty += 1
    density = round(nonempty / total_cells, 3) if total_cells else 0.0

    hints = []
    rows_with_multi_long = [
        r for r in long_cells_rows if long_cells_counts[str(r)] >= 2
    ]
    if rows_with_multi_long:
        hints.append("spec_block_above_data")
    widths_in_first_rows = [r["widths_eff"] for r in first_rows]
    for i in range(len(first_rows) - 1):
        if (
            widths_in_first_rows[i] >= 3
            and abs(widths_in_first_rows[i] - widths_in_first_rows[i + 1]) <= 1
            and first_rows[i]["merge_flag"] == "v-merge"
        ):
            hints.append("multi_tier_header_candidate")
            break
    if first_rows and first_rows[0]["widths_nb"] >= 3 and not long_cells_rows:
        hints.append("row1_db_candidate")
    if max_col <= 3 and long_cells_max_len >= 80:
        hints.append("prose_notes_candidate")
    if merges["label_col_merges"] >= 3 or (merges["count"] >= 20 and merges["banner_rows"]):
        hints.append("rules_table_candidate")

    return {
        "name": sheet_name,
        "max_row": max_row,
        "max_col": max_col,
        "nonempty_density": density,
        "first_rows": first_rows,
        "merges": merges,
        "long_cells": {
            "rows": long_cells_rows,
            "max_len": long_cells_max_len,
            "per_row_counts": long_cells_counts,
        },
        "deep_sample": {
            "median_width": median_width,
            "width_histogram": top_hist,
        },
        "structural_hints": hints,
    }


def probe(xlsx_path: str, sheet: Optional[str] = None) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    if sheet:
        return {"file": str(Path(xlsx_path).name), "sheets": [_probe_sheet(wb, sheet)]}
    out = {"file": str(Path(xlsx_path).name), "sheets": []}
    for sn in wb.sheetnames:
        out["sheets"].append(_probe_sheet(wb, sn))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--sheet", default=None)
    args = ap.parse_args()
    result = probe(args.xlsx, args.sheet)
    json.dump(result, sys.stdout, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
