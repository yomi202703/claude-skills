#!/usr/bin/env python3
"""xlsxの範囲指定読み＋結合情報付き出力（Claude Code用）。

メタを見て必要と判断した範囲だけ読むための本読みスクリプト。
結合情報を添えて返せるので、後段で forward-fill/見出しスパン/連結 の判断ができる。

使用例:
  python3 xlsx_read.py foo.xlsx --sheet "003申込書"
  python3 xlsx_read.py foo.xlsx --sheet "003申込書" --range A1:V2
  python3 xlsx_read.py foo.xlsx --sheet "チェックシート" --range A5:H30 --include-merges --format json
"""
import sys
import json
import argparse
from pathlib import Path
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter


def parse_args():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("file", help="xlsx path")
    p.add_argument("--sheet", help="sheet name (default: first sheet)")
    p.add_argument("--range", dest="range_", help="cell range e.g. A1:H20 (default: entire used range)")
    p.add_argument("--include-merges", action="store_true", help="include merged range info")
    p.add_argument("--format", choices=["tsv", "json"], default="tsv")
    return p.parse_args()


def main():
    a = parse_args()
    wb = openpyxl.load_workbook(a.file, data_only=True)
    sheet_name = a.sheet if a.sheet else wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        print(f"[error] sheet '{sheet_name}' not found. available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(2)
    ws = wb[sheet_name]

    if a.range_:
        min_col, min_row, max_col, max_row = range_boundaries(a.range_)
    else:
        min_col, min_row = 1, 1
        max_col, max_row = ws.max_column, ws.max_row

    rows = []
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True,
    ):
        rows.append(list(row))

    merges = []
    if a.include_merges:
        for r in ws.merged_cells.ranges:
            if r.max_row < min_row or r.min_row > max_row:
                continue
            if r.max_col < min_col or r.min_col > max_col:
                continue
            top_left_value = ws.cell(row=r.min_row, column=r.min_col).value
            merges.append({
                "range": str(r),
                "value": top_left_value,
                "rows": [r.min_row, r.max_row],
                "cols": [r.min_col, r.max_col],
            })

    if a.format == "json":
        out = {
            "sheet": ws.title,
            "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
            "rows": rows,
        }
        if a.include_merges:
            out["merges"] = merges
        print(json.dumps(out, ensure_ascii=False, indent=2, default=str))
    else:
        for row in rows:
            print("\t".join("" if v is None else str(v) for v in row))
        if a.include_merges and merges:
            print("", file=sys.stderr)
            print("--- merged ranges in output ---", file=sys.stderr)
            for m in merges:
                print(f"  {m['range']}\tvalue={m['value']}", file=sys.stderr)


if __name__ == "__main__":
    main()
