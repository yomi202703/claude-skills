#!/usr/bin/env python3
"""xlsxを分類し、各シートに適用すべきパス(P1-P5)と読むべきdocsを出力する。

SKILL.mdのルーター用スクリプト。この出力を見てClaudeはdocs/以下の
該当ドキュメントだけを読み、実行する。lazy-loadを実現する。

ヘッダ検出は2段式分類器:
  Stage 1 — pruning: バナー/【】/※/過疎行を除外 (knowledge/header_detection.md 準拠)
  Stage 2 — weighted scoring: 型フリップ (Adelfio&Samet VLDB'13) 主導 +
            formatting (Fan AAAI'12) + uniqueness + label-length

使用例:
  python3 xlsx_classify.py <file.xlsx>
"""
import sys
import json
from collections import Counter
from pathlib import Path
from typing import Optional
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import xlsx_primitives as _prim  # noqa: E402
from xlsx_drawings import extract as _extract_drawings  # noqa: E402


RULE_KEYWORDS = ["チェック", "確認", "ルール", "規定", "判定", "点検", "マニュアル"]
RULE_HEADERS = ["確認事項", "チェック欄", "判定基準", "条件"]
TABLE_KEYWORDS = ["一覧", "台帳", "マスタ", "リスト"]
DOC_KEYWORDS = ["仕様", "定義", "設計"]
NOTES_KEYWORDS = ["について", "ご案内", "お願い", "留意", "注意事項"]


def detect_notes_shape(ws, sample_limit: int = 500) -> bool:
    """True when the sheet is predominantly long-string notes (not tabular).

    Signals: most non-empty cells contain long strings, and most non-empty
    rows have only 1-2 non-empty cells. Samples up to `sample_limit` rows
    (notes sheets are usually small; we won't mis-classify a huge data table
    as notes because at large sizes the table-like patterns dominate).
    """
    long_cells = 0
    short_cells = 0
    row_counts = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx > sample_limit:
            break
        cnt = 0
        for v in row:
            if v is None:
                continue
            s = v if isinstance(v, str) else str(v)
            if not s.strip():
                continue
            cnt += 1
            if isinstance(v, str) and len(v.strip()) > 30:
                long_cells += 1
            else:
                short_cells += 1
        if cnt > 0:
            row_counts.append(cnt)
    if not row_counts:
        return False
    total_cells = long_cells + short_cells
    long_frac = long_cells / total_cells if total_cells else 0
    thin_row_frac = sum(1 for c in row_counts if c <= 2) / len(row_counts)
    return long_frac >= 0.5 and thin_row_frac >= 0.6


def infer_content_type(filename: str, sheet_name: str, header_preview: list, ws=None) -> str:
    haystack = " ".join([
        filename or "",
        sheet_name or "",
        " ".join(str(h) for h in header_preview if h is not None),
    ])
    # Notes sheets: keyword + shape-based detection
    if ws is not None and detect_notes_shape(ws):
        return "notes"
    for kw in NOTES_KEYWORDS:
        if kw in (sheet_name or ""):
            return "notes"
    for kw in RULE_HEADERS + RULE_KEYWORDS:
        if kw in haystack:
            return "rules"
    for kw in TABLE_KEYWORDS:
        if kw in haystack:
            return "table"
    for kw in DOC_KEYWORDS:
        if kw in haystack:
            return "document"
    return "table"


def _row_values(ws, row_i: int) -> list:
    return [ws.cell(row=row_i, column=c).value for c in range(1, ws.max_column + 1)]


def _sample_deep_row_width(ws, skip_top: int = 30, sample: int = 200) -> int:
    """Median number of non-empty cells in rows beyond `skip_top`.

    Used to characterize the "typical data row width" so header candidates
    can be penalized when their populated-cell count is far below it
    (e.g. spec-block rows embedded in a documentation prelude).
    Returns 0 when the sheet has fewer than `skip_top` rows.
    """
    widths = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx <= skip_top:
            continue
        if idx > skip_top + sample:
            break
        cnt = sum(1 for v in row
                  if v is not None and (not isinstance(v, str) or v.strip()))
        if cnt > 0:
            widths.append(cnt)
    if not widths:
        return 0
    widths.sort()
    return widths[len(widths) // 2]


def _score_header_candidate(ws, row_i: int, k: int = 3, typical_width: int = 0) -> tuple:
    """Stage-2 scoring. Returns (score, features_dict).

    Weights derived from knowledge/header_detection.md recommendations:
    type-flip is the dominant signal (Adelfio&Samet VLDB'13).
    """
    n_cols = ws.max_column
    r_vals = _row_values(ws, row_i)
    r_types = [_prim.cell_type(v) for v in r_vals]
    # Effective column count for fraction-based features. Using raw max_column
    # deflates str_frac/bold_frac/empty_frac on sheets with many sparse
    # trailing columns (e.g. max_column=42 but real data only ~8 cols wide).
    # Only kick in when trailing sparsity is obvious (max_col >> typical_width);
    # for normal DB sheets (max_col ≈ real width) keep n_cols unchanged so we
    # don't inflate scores for partial-width rows inside the data region.
    this_w = sum(1 for v in r_vals if v is not None and (not isinstance(v, str) or v.strip()))
    if typical_width >= 3 and n_cols > typical_width * 2:
        n_cols_eff = max(this_w, typical_width, 3)
    else:
        n_cols_eff = n_cols

    # Collect lookahead rows
    lookahead_types = []
    for d in range(1, k + 1):
        nr = row_i + d
        if nr > ws.max_row:
            break
        lookahead_types.append([_prim.cell_type(v) for v in _row_values(ws, nr)])

    # Feature: type-flip per column (R is string, below majority is non-string/non-empty)
    flip_cnt = 0
    eval_cols = 0
    if lookahead_types:
        for c in range(n_cols):
            below = [la[c] for la in lookahead_types if c < len(la) and la[c] != "e"]
            if not below:
                continue
            eval_cols += 1
            below_majority = Counter(below).most_common(1)[0][0]
            if r_types[c] == "s" and below_majority != "s":
                flip_cnt += 1
    flip_frac = flip_cnt / eval_cols if eval_cols else 0.0

    # Feature: string fraction in R (normalized by effective width)
    str_frac = sum(1 for t in r_types if t == "s") / max(1, n_cols_eff)
    str_frac = min(str_frac, 1.0)

    # Feature: formatting (bold) — only scan populated cells to match n_cols_eff
    bold_cnt = 0
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_i, column=c)
        if cell.font and cell.font.bold:
            bold_cnt += 1
    bold_frac = min(bold_cnt / max(1, n_cols_eff), 1.0)

    # Feature: uniqueness of non-empty string values
    non_empty_strs = [v.strip() for v in r_vals if isinstance(v, str) and v.strip()]
    uniq_frac = len(set(non_empty_strs)) / max(1, len(non_empty_strs)) if non_empty_strs else 0.0

    # Feature: mean string length → label-like. Broadened to not zero out
    # long descriptive-header rows that are still legitimate headers.
    if non_empty_strs:
        mean_len = sum(len(s) for s in non_empty_strs) / len(non_empty_strs)
        if mean_len <= 15:
            len_bonus = 1.0
        elif mean_len <= 30:
            len_bonus = 0.7
        elif mean_len <= 60:
            len_bonus = 0.4
        else:
            len_bonus = 0.0
    else:
        mean_len = 0
        len_bonus = 0.0

    # Feature: descriptive_contrast — header rows tend to have longer label
    # strings than the data rows immediately below. Rewards a row whose
    # mean string length is substantially greater than the following rows.
    # Critical when type-flip fails (e.g. code-column tables that are all-string).
    below_means = []
    for d in range(1, k + 1):
        nr = row_i + d
        if nr > ws.max_row:
            break
        below_vals = _row_values(ws, nr)
        below_strs = [v.strip() for v in below_vals if isinstance(v, str) and v.strip()]
        if below_strs:
            below_means.append(sum(len(s) for s in below_strs) / len(below_strs))
    below_mean = sum(below_means) / len(below_means) if below_means else 0.0
    desc_contrast = 0.0
    if mean_len >= 12 and below_mean >= 1:
        ratio = mean_len / max(1.0, below_mean)
        if ratio >= 3.0:
            desc_contrast = 2.0
        elif ratio >= 2.5:
            desc_contrast = 1.5
        elif ratio >= 1.8:
            desc_contrast = 1.0

    # Feature: post_spec_block — distance-weighted bonus when one of the 5
    # rows above is a spec/documentation block (≥2 long prose cells of ≥150
    # chars). Requiring ≥2 long cells avoids false positives from data rows
    # that happen to contain one long description column (e.g. drug-info rows
    # in a medical rules DB) — real spec blocks span multiple cells because
    # they're 概要/目的/注意事項-style multi-column documentation.
    post_spec_block = 0.0
    for d in range(1, 6):
        ar = row_i - d
        if ar < 1:
            break
        long_cells = sum(
            1 for v in _row_values(ws, ar)
            if isinstance(v, str) and len(v) >= 150
        )
        if long_cells >= 2:
            post_spec_block = max(0.0, 1.25 - 0.25 * (d - 1))
            break

    # Feature: empty fraction
    empty_frac = sum(1 for t in r_types if t == "e") / max(1, n_cols)

    # Feature: row inside vertical merge
    in_v_merge = any(
        r.min_row <= row_i <= r.max_row and r.min_row != r.max_row
        for r in ws.merged_cells.ranges
    )

    # Feature: matches typical data-row width (prevents picking sparse spec rows
    # in documentation preludes when the real data table has more columns)
    this_width = sum(1 for v in r_vals if v is not None and (not isinstance(v, str) or v.strip()))
    shape_match = 0.0
    if typical_width >= 3:
        if this_width >= typical_width * 0.7:
            shape_match = 1.0
        elif this_width < typical_width * 0.4:
            shape_match = -1.0

    # Weighted sum — flip_frac dominant, shape_match adds strong tiebreak
    score = (
        2.5 * flip_frac
        + 1.0 * str_frac
        + 1.0 * bold_frac
        + 0.5 * uniq_frac
        + 1.0 * len_bonus
        + 1.5 * shape_match
        + 1.2 * desc_contrast
        + 1.5 * post_spec_block
        - 0.5 * empty_frac
        - 0.5 * (1.0 if in_v_merge else 0.0)
    )
    return score, {
        "flip_frac": round(flip_frac, 2),
        "str_frac": round(str_frac, 2),
        "bold_frac": round(bold_frac, 2),
        "uniq_frac": round(uniq_frac, 2),
        "mean_len": round(mean_len, 1),
        "empty_frac": round(empty_frac, 2),
        "shape_match": shape_match,
        "desc_contrast": desc_contrast,
        "post_spec_block": post_spec_block,
        "in_v_merge": in_v_merge,
    }


def detect_header_row(ws, max_scan: int = 50) -> dict:
    """Two-stage header-row classifier.

    Stage 1 (rule-based pruning):
      - skip rows that are full-width horizontal banners (merge spanning ~all cols)
      - skip rows containing 【 or 】 anywhere
      - skip rows whose first non-empty cell starts with ※
      - skip rows with < 2 non-empty cells

    Stage 2 (weighted feature scoring):
      see _score_header_candidate — type-flip is dominant.

    Returns:
      {
        "primary": <row_index>,           # best header row
        "rows": [<r>] or [<r>, <r+1>],    # multi-tier if top 2 contiguous & both high
        "confidence": float,
        "features": dict,
      }
    """
    banner_rows = set()
    for r in ws.merged_cells.ranges:
        if r.min_row == r.max_row and (r.max_col - r.min_col + 1) >= max(3, ws.max_column - 1):
            banner_rows.add(r.min_row)

    limit = min(max_scan, ws.max_row)
    candidates = []
    for i in range(1, limit + 1):
        if i in banner_rows:
            continue
        row_vals = _row_values(ws, i)
        # Accept as candidate if either:
        #   (a) ≥3 short-label strings (≤30 chars) — classic simple header, OR
        #   (b) ≥5 medium strings (≤120 chars) — long descriptive header with \n
        short_strs = [v for v in row_vals
                      if isinstance(v, str) and v.strip() and len(v.strip()) <= 30]
        medium_strs = [v for v in row_vals
                       if isinstance(v, str) and v.strip() and len(v.strip()) <= 120]
        if len(short_strs) < 3 and len(medium_strs) < 5:
            continue
        all_strs = short_strs if len(short_strs) >= 3 else medium_strs
        if any(("【" in s or "】" in s) for s in all_strs):
            continue
        if any(s.lstrip().startswith("※") for s in all_strs):
            continue
        candidates.append(i)

    if not candidates:
        return {"primary": 1, "rows": [1], "confidence": 0.0, "features": {}}

    typical_width = _sample_deep_row_width(ws)

    # Pre-compute each row's non-empty width WITH forward-fill from merge anchors.
    # Without this, data rows under a merged label column (e.g. 区分 at A4:A30)
    # look "narrower" than the header, which breaks run-based heuristics.
    merge_lookup = _prim.build_merge_lookup(ws)
    widths = [
        _prim.row_effective_width(ws, r, merge_lookup)
        for r in range(1, min(limit + 500, ws.max_row) + 1)
    ]

    def run_start_bonus(row_i: int) -> float:
        """Bonus = (forward_run − backward_run_within_tol) * 0.1, capped at 2.0.
        Rewards rows that start a consistent-width region WITHOUT already being
        inside a similar region above. A header matches its data width within
        ±tol, so we can't require sharp delta, but mid-sheet data rows get
        canceled out by their own backward run.
        """
        w_here = widths[row_i - 1] if row_i - 1 < len(widths) else 0
        if w_here < 3:
            return 0.0
        tol = max(1, int(w_here * 0.25))
        # forward: rows below at width within tol
        fwd = 0
        for j in range(row_i, len(widths)):
            if widths[j] == 0:
                break
            if abs(widths[j] - w_here) <= tol:
                fwd += 1
            else:
                break
            if fwd >= 20:
                break
        # backward: rows above at width within tol (how deep we're already in a run)
        back = 0
        for j in range(row_i - 2, -1, -1):
            if widths[j] == 0:
                break
            if abs(widths[j] - w_here) <= tol:
                back += 1
            else:
                break
            if back >= 20:
                break
        net = max(fwd - back, 0)
        return min(net, 20) * 0.1

    scored = []
    for i in candidates:
        score, features = _score_header_candidate(ws, i, typical_width=typical_width)
        row_vals = _row_values(ws, i)
        # Short-label bonus — rewards header-like rows with many short unique labels.
        # Critical for simple DB sheets where row 1 is a plain label header
        # (type-flip and descriptive_contrast both fail there).
        short_cnt = sum(1 for v in row_vals
                        if isinstance(v, str) and v.strip() and len(v.strip()) <= 15)
        score += 0.3 * min(short_cnt, 5) * features["uniq_frac"]
        rsb = run_start_bonus(i)
        score += rsb
        features["short_cnt"] = short_cnt
        features["typical_width"] = typical_width
        features["run_start_bonus"] = round(rsb, 2)
        scored.append((score, i, features))
    scored.sort(key=lambda x: (-x[0], x[1]))  # high score first, earliest row for ties

    top_score, top_row, top_features = scored[0]
    rows = [top_row]

    # Multi-tier: parent-child header hierarchy.
    # Require: top 2 contiguous, second ≥ 75% of top, top score ≥ 1.5,
    #          AND the earlier of the two has ≥1 horizontal merge spanning >1 non-banner col.
    if len(scored) >= 2 and top_score >= 1.5:
        second_score, second_row, _ = scored[1]
        if abs(top_row - second_row) == 1 and second_score >= top_score * 0.75:
            parent_row = min(top_row, second_row)
            has_hierarchy_merge = any(
                r.min_row == parent_row and r.max_row == parent_row
                and 2 <= (r.max_col - r.min_col + 1) < ws.max_column - 1
                for r in ws.merged_cells.ranges
            )
            if has_hierarchy_merge:
                rows = sorted([top_row, second_row])

    return {
        "primary": top_row,
        "rows": rows,
        "confidence": round(top_score, 2),
        "features": top_features,
    }


def detect_data_rows(ws, header_row: int) -> list:
    """Estimate the first and last data row.

    First = header_row + 1. Last = the last row that has any non-None cell.
    Single-pass via iter_rows(values_only=True) for speed on very large sheets.
    """
    first = header_row + 1
    last = first - 1
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx < first:
            continue
        if any(v is not None and (not isinstance(v, str) or v.strip()) for v in row):
            last = idx
    return [first, max(first, last)]


def _coarse_type(v) -> Optional[str]:
    """Coarse cell type for orientation analysis. None for empty cells.

    All-digit strings collapse to 'n' so that text-formatted numbers
    (zip/phone stored as text) don't spuriously split a row/column that
    is otherwise numeric. Everything else textual is 's'.
    """
    import datetime as _dt
    if v is None:
        return None
    if isinstance(v, bool):
        return "n"
    if isinstance(v, (int, float)):
        return "n"
    if isinstance(v, (_dt.datetime, _dt.date)):
        return "d"
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        if s.isdigit():
            return "n"
        return "s"
    return "s"


def detect_transposed(ws, max_rows: int = 60, max_cols: int = 16) -> dict:
    """Detect a field-major (transposed) layout: rows are fields, columns are records.

    Principle: in a normal table each COLUMN is type-homogeneous and rows
    vary; in a transposed table each ROW is type-homogeneous and columns
    vary. We sample a window, fold text-formatted numbers, drop leading
    all-text label columns, then compare row-wise vs column-wise type
    consistency. A symmetric numeric matrix scores high on both axes and
    is intentionally NOT flagged (difference ≈ 0).

    Returns {transposed, confidence, row_consistency, col_consistency,
    label_cols}. Cheap: O(max_rows * max_cols), no extra file I/O.
    """
    nrows = min(ws.max_row, max_rows)
    ncols = min(ws.max_column, max_cols)
    null = {"transposed": False, "confidence": 0.0,
            "row_consistency": 0.0, "col_consistency": 0.0, "label_cols": 0}
    if nrows < 5 or ncols < 4:
        return null

    grid = [[_coarse_type(ws.cell(row=r, column=c).value)
             for c in range(1, ncols + 1)]
            for r in range(1, nrows + 1)]

    # Drop leading label columns: among the first 3 columns, a column whose
    # non-empty cells are predominantly text ('s') is a label spine, not data.
    label_cols = 0
    for c in range(min(3, ncols)):
        col = [grid[r][c] for r in range(nrows) if grid[r][c] is not None]
        if len(col) >= 3 and sum(t == "s" for t in col) / len(col) >= 0.7:
            label_cols += 1
        else:
            break
    data_cols = list(range(label_cols, ncols))
    if len(data_cols) < 3:
        return null

    def consistency(types: list) -> Optional[float]:
        nonempty = [t for t in types if t is not None]
        if len(nonempty) < 3:
            return None
        modal = Counter(nonempty).most_common(1)[0][1]
        return modal / len(nonempty)

    row_scores = [consistency([grid[r][c] for c in data_cols]) for r in range(nrows)]
    col_scores = [consistency([grid[r][c] for r in range(nrows)]) for c in data_cols]
    row_scores = [s for s in row_scores if s is not None]
    col_scores = [s for s in col_scores if s is not None]
    if len(row_scores) < 4 or not col_scores:
        return null

    row_consistency = sum(row_scores) / len(row_scores)
    col_consistency = sum(col_scores) / len(col_scores)
    diff = row_consistency - col_consistency
    transposed = row_consistency >= 0.75 and diff >= 0.20
    return {
        "transposed": transposed,
        "confidence": round(diff, 3),
        "row_consistency": round(row_consistency, 3),
        "col_consistency": round(col_consistency, 3),
        "label_cols": label_cols,
    }


def classify_sheet(wb, sn: str, filename: str, drawings_info: Optional[dict] = None) -> dict:
    ws = wb[sn]
    merges = list(ws.merged_cells.ranges)
    header_info = detect_header_row(ws)
    header_row_index = header_info["primary"]
    header_rows = header_info["rows"]
    header_row = [ws.cell(row=header_row_index, column=c).value for c in range(1, ws.max_column + 1)]
    # If multi-tier header, data starts AFTER the last header row
    last_header_row = max(header_rows)
    data_rows_range = detect_data_rows(ws, last_header_row)
    data_row_count = max(0, data_rows_range[1] - data_rows_range[0] + 1)

    looks_like_db = (
        len(merges) == 0
        and all(v is not None for v in header_row[: min(5, len(header_row))])
    )
    shape = "db" if looks_like_db else "structured"
    rows = ws.max_row
    cols = ws.max_column

    # Field-major (transposed) detection. When true, the cell-based header row
    # is meaningless — rows are fields, columns are records — so downstream must
    # follow transposed.md instead of treating header_row_index as a header.
    transpose_info = detect_transposed(ws)

    content_type = infer_content_type(filename, sn, header_row[:10], ws=ws)
    # Low header confidence ⇒ there's no real table structure here (flow
    # diagrams, prose, nested notes). Route to P3 notes rather than forcing
    # a table and emitting col_N columns that a downstream LLM can't use.
    if header_info["confidence"] < 1.5 and content_type == "table":
        content_type = "notes"

    # Size is driven by actual data size, not padded max_row. P5 reserved for
    # truly prohibitive sheets (>500k cells OR >50k data rows OR >200 cols).
    effective_cells = data_row_count * cols
    if effective_cells > 500_000 or data_row_count > 50_000 or cols > 200:
        size = "huge"
        path = "P5"
    elif content_type == "notes":
        size = "small" if data_row_count <= 100 else "large"
        path = "P3" if size == "small" else "P4"
    elif shape == "db":
        size = "small" if data_row_count <= 100 else "large"
        path = "P1" if size == "small" else "P2"
    else:
        size = "small" if data_row_count <= 50 else "large"
        path = "P3" if size == "small" else "P4"

    # Drawing signals (Tier A). drawings_info comes from xlsx_drawings.extract()
    # for this sheet (or None when the caller didn't supply it). Keep zero-filled
    # fields for structural stability across outputs.
    d = drawings_info or {}
    d_counts = d.get("counts") or {}
    shape_n = int(d_counts.get("shape", 0))
    pic_n = int(d_counts.get("pic", 0))
    cxn_n = int(d_counts.get("cxn", 0))
    total_draw = shape_n + pic_n + cxn_n
    # has_drawings reflects ACTUAL extracted content (shape/pic/cxn), not the
    # mere presence of a drawing XML part — many workbooks carry empty wsDr
    # shells (leftover from charts / template boilerplate) that should not
    # trigger the drawings pipeline.
    has_drawings = total_draw > 0
    has_vml = bool(d.get("has_vml_drawings"))
    drawing_density = round(total_draw / max(1, data_row_count), 3)
    # Visual-path heuristic: drawings likely carry semantic content.
    # - many pictures (product diagrams, flowcharts) OR
    # - many shapes with text (annotations over a layout) OR
    # - drawings exist AND the cell-based header is weak (content lives in drawings)
    suggests_visual = (
        pic_n >= 3
        or shape_n >= 5
        or (has_drawings and header_info["confidence"] < 1.0)
    )
    drawing_docs = ["drawings.md"] if has_drawings else []

    docs_map = {
        "P1": ["p1_db_small.md"],
        "P2": ["p2_db_large.md"],
        "P3": ["p3_structured_small.md", "merges.md"],
        "P4": ["p4_structured_large.md", "merges.md"],
        "P5": ["p5_metadata_only.md"],
    }
    sheet_docs = docs_map[path] + drawing_docs
    if suggests_visual:
        sheet_docs.append("p6_visual.md")
    if transpose_info["transposed"]:
        sheet_docs.append("transposed.md")

    return {
        "name": sn,
        "sheet_slug": sanitize_slug(sn),
        "rows": rows,
        "cols": cols,
        "merged_count": len(merges),
        "shape": shape,
        "size": size,
        "content_type": content_type,
        "path": path,
        "docs_to_read": sheet_docs,
        "header_row_index": header_row_index,
        "header_rows": header_rows,
        "header_confidence": header_info["confidence"],
        "header_features": header_info["features"],
        "data_rows": data_rows_range,
        "data_row_count": data_row_count,
        "header_preview": header_row[:10],
        "transposed": transpose_info["transposed"],
        "transpose_confidence": transpose_info["confidence"],
        "has_drawings": has_drawings,
        "has_vml_drawings": has_vml,
        "shape_count": shape_n,
        "pic_count": pic_n,
        "cxn_count": cxn_n,
        "drawing_density": drawing_density,
        "suggests_visual": suggests_visual,
    }


sanitize_slug = _prim.sanitize_slug  # re-export for backward compat


def main():
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        sys.exit(2)

    for path_arg in sys.argv[1:]:
        p = Path(path_arg).resolve()
        wb = openpyxl.load_workbook(p, data_only=True)
        filename = p.name
        # Extract drawings once per workbook; pass per-sheet info into the classifier.
        try:
            drawings_all = _extract_drawings(p)
            drawings_by_sheet = drawings_all.get("sheets", {})
        except Exception:
            drawings_by_sheet = {}
        sheets = [
            classify_sheet(wb, sn, filename, drawings_info=drawings_by_sheet.get(sn))
            for sn in wb.sheetnames
        ]
        multi_sheet = len(sheets) > 1
        extra_docs = ["multi_sheet.md", "manifest.md"] if multi_sheet else []
        workbook_slug = sanitize_slug(p.stem)

        all_docs = set()
        for s in sheets:
            all_docs.update(s["docs_to_read"])
        all_docs.update(extra_docs)

        out = {
            "file": str(p),
            "filename": filename,
            "workbook_slug": workbook_slug,
            "multi_sheet": multi_sheet,
            "sheets": sheets,
            "extra_docs": extra_docs,
            "all_docs_to_read": sorted(all_docs),
            "output_dir_suggestion": f"data/{workbook_slug}",
            "docs_base_dir": "~/.claude/skills/xlsx-router/docs",
        }
        print(json.dumps(out, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
